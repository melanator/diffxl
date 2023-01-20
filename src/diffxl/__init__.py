import argparse
from pathlib import Path
from openpyxl import load_workbook, Workbook
from outformat import *


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument("file_1", help="First excel file")
    parser.add_argument("file_2", help="Second excel file")
    parser.add_argument("-v", "--verbose", help="Verbose mode", action="store_true")
    args = parser.parse_args()

    file1, file2 = Path(args.file_1), Path(args.file_2)

    # Check if first path exists
    if not file1.exists():
        print("The first target file does not exists")
        raise SystemExit(1)
    
    # Check if second path exists
    if not file2.exists():
        print("The second target file does not exists")
        raise SystemExit(1)
    
    # Check if path are the same
    if file1 == file2:
        print("You selected the same workbooks\nExit")
        return SystemExit(2)

    compare(load_workbook(file1), load_workbook(file2))
   

def compare(wb1: Workbook, wb2: Workbook):  
    if(wb1 == wb2):
        print("Workbooks are the same")
        return 0
    
    # Naive traverse over all sheets
    for sheet in wb1.sheetnames:
        
        ws = wb1[sheet]
        
        if sheet not in wb2.sheetnames:
            print(RedString(f"Sheet {sheet} is deleted"))
        else:
            print("Changes for sheet:", RedString(sheet))
            count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value != wb2[sheet][cell.coordinate].value:
                        other_cell = wb2[sheet][cell.coordinate].value
                        count += 1
                        if cell.value and other_cell:
                            print(YellowString(f"{cell.coordinate}:\t {cell.value} -> {other_cell}"))
                        elif wb2[sheet][cell.coordinate].value:
                            print(GreenString(f"{cell.coordinate}:\t {other_cell}"), "deleted")
                        else:
                            print(RedString(f"{cell.coordinate}:\t {cell.value}"), "added")
                        

        if count == 0:
            print(GreenString("Sheet has no changes"))
            print("-"*20)
    
    for sheet in wb2.sheetnames:
        if sheet not in wb1.sheetnames:
            print(GreenString(f"Sheet {sheet} is deleted"))
        

if __name__ == "__main__":
    main()
